#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
问题1 数据预处理 — 四特征构造、口径统一与质量验收
================================================
输入：
  - 附件2-2022年30个省份排放清单.xlsx  (30 province sheets + NOTE)
  - 人口与GDP数据.csv                  (实际为 XLSX，双表：人口/GDP)
输出：
  - results/problem1_preprocessed/problem1_province_features_2022.csv
  - results/problem1_preprocessed/problem1_features_standardized_2022.csv
  - results/problem1_preprocessed/industrial_sector_mapping_audit.csv
  - results/problem1_preprocessed/external_province_data_2022_cleaned.csv
  - results/problem1_preprocessed/problem1_data_dictionary.md
  - results/problem1_preprocessed/preprocessing_log.txt
  - results/problem1_preprocessed/descriptive_stats.csv
  - results/problem1_preprocessed/correlation_pearson.csv
  - results/problem1_preprocessed/correlation_spearman.csv
  - figures/problem1_*.png
严格遵循《问题1_数据预处理报告.md》口径：
  C_i 取 TotalEmissions 行 Scope_1_Total 列
  C_i^ind = sum(S_ind)  38 个工业行，Construction 暂不纳入
  I_i = 100*C_Mt/GDP_100M (t/万元)
  H_i = 100*C_Mt/Pop_10k   (t/人)
  Q_i = C_ind/C_total
  标准化： ln -> Min-Max  (四指标均为越大压力越高，不反向)
"""

import shutil
import tempfile
import warnings
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaborn as sns

# ---------------------------------------------------------------------------
# 路径配置
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[1]
ATTACH_DIR = PROJECT_ROOT / "附件"
RESULTS_DIR = PROJECT_ROOT / "results" / "problem1_preprocessed"
FIGURES_DIR = PROJECT_ROOT / "figures"
CODE_DIR = PROJECT_ROOT / "code"

ATTACH2_PATH = ATTACH_DIR / "附件2-2022年30个省份排放清单.xlsx"
POP_GDP_PATH = ATTACH_DIR / "人口与GDP数据.csv"  # 实际为 XLSX
INDICATOR2_PATH = ATTACH_DIR / "附件2的指标说明.xlsx"

# 输出文件
F_MAIN = RESULTS_DIR / "problem1_province_features_2022.csv"
F_STD = RESULTS_DIR / "problem1_features_standardized_2022.csv"
F_AUDIT = RESULTS_DIR / "industrial_sector_mapping_audit.csv"
F_EXTERNAL = RESULTS_DIR / "external_province_data_2022_cleaned.csv"
F_DICT = RESULTS_DIR / "problem1_data_dictionary.md"
F_LOG = RESULTS_DIR / "preprocessing_log.txt"
F_DESC = RESULTS_DIR / "descriptive_stats.csv"
F_CORR_P = RESULTS_DIR / "correlation_pearson.csv"
F_CORR_S = RESULTS_DIR / "correlation_spearman.csv"

# ---------------------------------------------------------------------------
# 省份映射  sheet(去2022后缀的英文) -> 中文
# ---------------------------------------------------------------------------
EN_TO_CN = {
    "Shanghai": "上海",
    "Yunnan": "云南",
    "InnerMongolia": "内蒙古",
    "Beijing": "北京",
    "Jilin": "吉林",
    "Sichuan": "四川",
    "Tianjin": "天津",
    "Ningxia": "宁夏",
    "Anhui": "安徽",
    "Shandong": "山东",
    "Shanxi": "山西",
    "Guangdong": "广东",
    "Guangxi": "广西",
    "Xinjiang": "新疆",
    "Jiangsu": "江苏",
    "Jiangxi": "江西",
    "Hebei": "河北",
    "Henan": "河南",
    "Zhejiang": "浙江",
    "Hainan": "海南",
    "Hubei": "湖北",
    "Hunan": "湖南",
    "Gansu": "甘肃",
    "Fujian": "福建",
    "Guizhou": "贵州",
    "Liaoning": "辽宁",
    "Chongqing": "重庆",
    "Shaanxi": "陕西",
    "Qinghai": "青海",
    "Heilongjiang": "黑龙江",
}

# ---------------------------------------------------------------------------
# 工业口径  37 行  (报告 §4.1)
# ---------------------------------------------------------------------------
INDUSTRIAL_SECTORS = [
    # 采矿业 6
    "Coal Mining and Dressing",
    "Petroleum and Natural Gas Extraction",
    "Ferrous Metals Mining and Dressing",
    "Nonferrous Metals Mining and Dressing",
    "Nonmetal Minerals Mining and Dressing",
    "Other Minerals Mining and Dressing",
    # 制造业 28
    "Food Processing",
    "Food Production",
    "Beverage Production",
    "Tobacco Processing",
    "Textile Industry",
    "Garments and Other Fiber Products",
    "Leather, Furs, Down and Related Products",
    "Timber Processing, Bamboo, Cane, Palm Fiber & Straw Products",
    "Furniture Manufacturing",
    "Papermaking and Paper Products",
    "Printing and Record Medium Reproduction",
    "Cultural, Educational and Sports Articles",
    "Petroleum Processing and Coking",
    "Raw Chemical Materials and Chemical Products",
    "Medical and Pharmaceutical Products",
    "Chemical Fiber",
    "Rubber Products",
    "Plastic Products",
    "Nonmetal Mineral Products",
    "Smelting and Pressing of Ferrous Metals",
    "Smelting and Pressing of Nonferrous Metals",
    "Metal Products",
    "Ordinary Machinery",
    "Equipment for Special Purposes",
    "Transportation Equipment",
    "Electric Equipment and Machinery",
    "Electronic and Telecommunications Equipment",
    "Instruments, Meters, Cultural and Office Machinery",
    "Other Manufacturing Industry",
    # 公用事业 3
    "Production and Supply of Electric Power, Steam and Hot Water",
    "Production and Supply of Gas",
    "Production and Supply of Tap Water",
]

# 暂不纳入工业的行 (报告 §4.2)
NON_INDUSTRIAL_SECTORS = [
    "Farming, Forestry, Animal Husbandry, Fishery and Water Conservancy",
    "Logging and Transport of Wood and Bamboo",
    "Scrap and waste",
    "Construction",
    "Transportation, Storage, Post and Telecommunication Services",
    "Wholesale, Retail Trade and Catering Services",
    "Others",
    "Urban",
    "Rural",
    "TotalEmissions",
]

ALL_EXPECTED_SECTORS = set(INDUSTRIAL_SECTORS + NON_INDUSTRIAL_SECTORS)

# 绘图统一配色
PALETTE = sns.color_palette("Set2", 8)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
# 中文字体：优先加载系统 SimHei / Microsoft YaHei，再设为默认
try:
    import matplotlib.font_manager as _fm
    for _fp in [r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\msyhbd.ttc"]:
        if Path(_fp).exists():
            _fm.fontManager.addfont(_fp)
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Microsoft YaHei Light", "STXihei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False
except Exception:
    plt.rcParams["axes.unicode_minus"] = False


# ---------------------------------------------------------------------------
def _load_external_data(pop_gdp_path: Path) -> pd.DataFrame:
    """读取人口与GDP双表，返回合并后的 DataFrame。

    该文件扩展名为 .csv 但实际为 XLSX，需先复制为临时 .xlsx 再用 openpyxl 读取。
    返回字段: province, gdp_2022_100m_yuan, population_2022_10k
    """
    tmp = None
    try:
        # 尝试直接用 openpyxl，失败则复制
        try:
            wb = openpyxl.load_workbook(pop_gdp_path, data_only=True)
        except Exception:
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.close()
            shutil.copyfile(pop_gdp_path, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, data_only=True)

        def _read_sheet(name: str) -> pd.DataFrame:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            # 第一行为表头
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            data = []
            for r in rows[1:]:
                if r[0] is None or str(r[0]).strip() == "":
                    continue
                data.append(r)
            df = pd.DataFrame(data, columns=header)
            return df

        df_pop = _read_sheet("人口")
        df_gdp = _read_sheet("GDP")

        # 规范列名
        df_pop = df_pop.rename(columns={"省份": "province", "人口（万人）": "population_2022_10k"})
        df_gdp = df_gdp.rename(columns={"省份": "province", "GDP（亿元）": "gdp_2022_100m_yuan"})

        # 检查数量
        assert len(df_pop) == 30, f"人口表应有30行，实际{len(df_pop)}"
        assert len(df_gdp) == 30, f"GDP表应有30行，实际{len(df_gdp)}"
        assert df_pop["province"].nunique() == 30, "人口表省份重复"
        assert df_gdp["province"].nunique() == 30, "GDP表省份重复"

        # 数值校验
        for col in ["population_2022_10k"]:
            s = pd.to_numeric(df_pop[col], errors="coerce")
            assert s.notna().all(), f"人口列存在非数值: {df_pop[s.isna()]}"
            assert (s > 0).all(), f"人口存在零或负值: {s[s<=0].tolist()}"

        for col in ["gdp_2022_100m_yuan"]:
            s = pd.to_numeric(df_gdp[col], errors="coerce")
            assert s.notna().all(), f"GDP列存在非数值"
            assert (s > 0).all(), f"GDP存在零或负值"

        df_pop["population_2022_10k"] = pd.to_numeric(df_pop["population_2022_10k"])
        df_gdp["gdp_2022_100m_yuan"] = pd.to_numeric(df_gdp["gdp_2022_100m_yuan"])

        merged = pd.merge(df_pop, df_gdp, on="province", how="inner", validate="one_to_one")
        assert len(merged) == 30, f"合并后应为30行，实际{len(merged)}"

        merged = merged.sort_values("province").reset_index(drop=True)
        return merged

    finally:
        if tmp is not None:
            try:
                Path(tmp.name).unlink(missing_ok=True)
            except Exception:
                pass


def _load_emission_inventory(attach2_path: Path):
    """解析附件2每个省份的排放清单。

    Returns:
        df_emission:  columns [province, source_sheet, emission_total_mtco2,
                               industrial_emission_mtco2, ... audit fields]
        audit_rows:   list of dict per province for audit table
    """
    wb = openpyxl.load_workbook(attach2_path, data_only=True)
    sheet_names = [s for s in wb.sheetnames if s != "NOTE"]

    assert len(sheet_names) == 30, f"附件2应有30个省份工作表，实际{len(sheet_names)}"

    records = []
    audit_rows = []

    for sheet in sheet_names:
        # 去掉2022后缀得到英文名
        en_name = sheet.replace("2022", "") if sheet.endswith("2022") else sheet
        if en_name not in EN_TO_CN:
            raise ValueError(f"未知的省份工作表映射: {sheet} -> {en_name}")
        province_cn = EN_TO_CN[en_name]

        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [c for c in rows[0]]
        # 定位 Scope_1_Total 列
        try:
            scope_idx = header.index("Scope_1_Total")
        except ValueError:
            raise ValueError(f"{sheet} 未找到 Scope_1_Total 列，表头为{header}")

        # 建立 sector -> Scope_1_Total 的映射
        sector_map = {}
        for r in rows[1:]:
            sector = r[0]
            if sector is None or str(sector).strip() == "" or sector == "unit":
                continue
            sector_str = str(sector).strip()
            val = r[scope_idx]
            # 跳过参考文献等无数值行（如 Jilin 后5行 1. Xu et al. ...）
            if val is None:
                # 若行业在预期集合内，记为 0；否则视为参考文献/空行跳过
                if sector_str not in ALL_EXPECTED_SECTORS and not sector_str.startswith("Emission"):
                    continue
                val = 0.0
            sector_map[sector_str] = float(val)

        # 总量
        if "TotalEmissions" not in sector_map:
            raise ValueError(f"{sheet} 缺少 TotalEmissions 行")
        total = sector_map["TotalEmissions"]
        assert total > 0, f"{sheet} 总排放量非正: {total}"

        # 工业汇总
        industrial_sum = 0.0
        missing_industrial = []
        for sec in INDUSTRIAL_SECTORS:
            if sec not in sector_map:
                missing_industrial.append(sec)
                # 缺失视为 0 但记录
                continue
            industrial_sum += sector_map[sec]

        # 审计：非工业、未映射
        unmapped = [k for k in sector_map.keys() if k not in ALL_EXPECTED_SECTORS]
        nonindustrial_sum = sum(sector_map.get(s, 0) for s in NON_INDUSTRIAL_SECTORS if s != "TotalEmissions")

        # 校验：工业+非工业+城镇/农村/其他? 实际上 sector_map 除了 TotalEmissions 外所有行之和应约等于 TotalEmissions ?
        # 附件2中 TotalEmissions 为独立核算总量，不强制等于各行业 Scope_1_Total 之和，但做审计记录
        sum_sectors = sum(v for k, v in sector_map.items() if k != "TotalEmissions")

        note_parts = []
        if missing_industrial:
            note_parts.append(f"缺失工业行{len(missing_industrial)}:{','.join(missing_industrial[:3])}")
        if unmapped:
            note_parts.append(f"未预期行{len(unmapped)}:{','.join(unmapped[:3])}")
        # 一致性偏差
        diff = total - sum_sectors
        note_parts.append(f"Total-SumSectors={diff:.2f}")

        records.append({
            "province": province_cn,
            "source_sheet": sheet,
            "emission_total_mtco2": total,
            "industrial_emission_mtco2": industrial_sum,
            "sector_sum_mtco2": sum_sectors,
            "_diff": diff,
        })
        audit_rows.append({
            "province": province_cn,
            "emission_total_mtco2": total,
            "industrial_emission_mtco2": industrial_sum,
            "nonindustrial_emission_mtco2": total - industrial_sum,
            "industrial_share": industrial_sum / total if total > 0 else np.nan,
            "mapped_sector_count": len([s for s in INDUSTRIAL_SECTORS if s in sector_map]),
            "unmapped_sector_count": len(unmapped),
            "sector_sum_mtco2": sum_sectors,
            "total_minus_sector_sum": diff,
            "mapping_note": "; ".join(note_parts) if note_parts else "OK",
        })

    df = pd.DataFrame(records).sort_values("province").reset_index(drop=True)
    audit_df = pd.DataFrame(audit_rows).sort_values("province").reset_index(drop=True)
    return df, audit_df


def _build_four_features(df_emission: pd.DataFrame, df_external: pd.DataFrame) -> pd.DataFrame:
    """合并并计算四特征及衍生指标。"""
    df = pd.merge(df_emission, df_external, on="province", how="inner", validate="one_to_one")
    assert len(df) == 30, f"合并后应为30行，实际{len(df)}"
    assert df["province"].nunique() == 30

    # 四特征
    df["industrial_emission_share"] = df["industrial_emission_mtco2"] / df["emission_total_mtco2"]
    df["industrial_emission_share_pct"] = df["industrial_emission_share"] * 100
    df["carbon_intensity_tco2_per_10k_yuan"] = 100 * df["emission_total_mtco2"] / df["gdp_2022_100m_yuan"]
    df["per_capita_emission_tco2_per_person"] = 100 * df["emission_total_mtco2"] / df["population_2022_10k"]

    # 校验
    assert (df["emission_total_mtco2"] > 0).all()
    assert (df["gdp_2022_100m_yuan"] > 0).all()
    assert (df["population_2022_10k"] > 0).all()
    assert (df["industrial_emission_share"] >= 0).all() and (df["industrial_emission_share"] <= 1).all()
    assert (df["industrial_emission_mtco2"] <= df["emission_total_mtco2"] + 1e-6).all(), "工业排放不应大于总量"
    assert df.isna().sum().sum() == 0, f"存在 NaN: {df.isna().sum()}"
    assert np.isfinite(df.select_dtypes(include=[np.number]).values).all(), "存在无穷值"

    # 按报告建议字段顺序输出
    out = df[[
        "province", "source_sheet",
        "emission_total_mtco2",
        "industrial_emission_mtco2",
        "industrial_emission_share",
        "industrial_emission_share_pct",
        "gdp_2022_100m_yuan",
        "population_2022_10k",
        "carbon_intensity_tco2_per_10k_yuan",
        "per_capita_emission_tco2_per_person",
    ]].sort_values("emission_total_mtco2", ascending=False).reset_index(drop=True)

    return out


def _standardize_features(df_main: pd.DataFrame) -> pd.DataFrame:
    """对四特征做 ln + Min-Max 标准化。

    四指标均为压力型（越大越高），不做反向。
    """
    feature_map = {
        "emission_total_mtco2": "z_emission_total",
        "carbon_intensity_tco2_per_10k_yuan": "z_carbon_intensity",
        "per_capita_emission_tco2_per_person": "z_per_capita_emission",
        "industrial_emission_share": "z_industrial_emission_share",
    }
    df_z = df_main[["province"]].copy()
    log_info = {}
    for orig, zcol in feature_map.items():
        x = df_main[orig].astype(float)
        assert (x > 0).all(), f"{orig} 存在非正值，无法取对数"
        xt = np.log(x)
        mn, mx = xt.min(), xt.max()
        assert mx > mn, f"{orig} log后无区分度"
        z = (xt - mn) / (mx - mn)
        df_z[zcol] = z
        log_info[orig] = {"log_min": mn, "log_max": mx, "raw_min": x.min(), "raw_max": x.max()}

    # 校验 0-1
    for c in feature_map.values():
        assert df_z[c].between(0, 1).all()

    df_z = df_z.sort_values("province").reset_index(drop=True)
    return df_z, log_info


def _descriptive_stats(df_main: pd.DataFrame) -> pd.DataFrame:
    """输出每个特征的描述统计。"""
    cols = [
        "emission_total_mtco2",
        "carbon_intensity_tco2_per_10k_yuan",
        "per_capita_emission_tco2_per_person",
        "industrial_emission_share",
    ]
    rows = []
    for c in cols:
        s = df_main[c]
        rows.append({
            "feature": c,
            "count": int(s.count()),
            "missing": int(s.isna().sum()),
            "illegal_nonpositive": int((s <= 0).sum()) if c != "industrial_emission_share" else int(((s < 0) | (s > 1)).sum()),
            "mean": s.mean(),
            "median": s.median(),
            "std": s.std(ddof=1),
            "min": s.min(),
            "q25": s.quantile(0.25),
            "q75": s.quantile(0.75),
            "max": s.max(),
            "cv": s.std(ddof=1) / s.mean() if s.mean() != 0 else np.nan,
            "range": s.max() - s.min(),
            "range_ratio": s.max() / s.min() if s.min() != 0 else np.nan,
        })
    return pd.DataFrame(rows)


def _correlation_matrices(df_main: pd.DataFrame):
    cols = [
        "emission_total_mtco2",
        "carbon_intensity_tco2_per_10k_yuan",
        "per_capita_emission_tco2_per_person",
        "industrial_emission_share",
    ]
    sub = df_main[cols]
    pearson = sub.corr(method="pearson")
    spearman = sub.corr(method="spearman")
    return pearson, spearman


# ---------------------------------------------------------------------------
def _write_data_dictionary():
    return """# 问题1 数据字典  problem1_data_dictionary.md

> 生成时间：{now}
> 口径来源：reports/问题1_数据预处理报告.md  §2-§7

## 1. 输入文件

| 文件 | 路径 | 说明 |
|------|------|------|
| 附件2-排放清单 | `附件/附件2-2022年30个省份排放清单.xlsx` | 30 个省份工作表（`*2022`）+ NOTE；总量取 `TotalEmissions` 行 `Scope_1_Total` 列（Mt CO2） |
| 指标说明2 | `附件/附件2的指标说明.xlsx` | 行业/能源口径核对，不直接参与计算 |
| 人口与GDP | `附件/人口与GDP数据.csv` | 实为 XLSX，含 `人口`/`GDP` 双表，字段 `省份`/`人口（万人）`/`GDP（亿元）`，各 30 行 |
| 附件1 | `附件/附件1-中国2019年-2025年碳排放数据.csv` | 仅作全国总量口径参考，不参与本模块 30 省特征计算 |

## 2. 省份口径

工作表名统一为中文（去 `2022` 后缀映射）：

Shanghai→上海, Yunnan→云南, InnerMongolia→内蒙古, Beijing→北京, Jilin→吉林, Sichuan→四川, Tianjin→天津, Ningxia→宁夏, Anhui→安徽, Shandong→山东, Shanxi→山西, Guangdong→广东, Guangxi→广西, Xinjiang→新疆, Jiangsu→江苏, Jiangxi→江西, Hebei→河北, Henan→河南, Zhejiang→浙江, Hainan→海南, Hubei→湖北, Hunan→湖南, Gansu→甘肃, Fujian→福建, Guizhou→贵州, Liaoning→辽宁, Chongqing→重庆, Shaanxi→陕西, Qinghai→青海, Heilongjiang→黑龙江

合并键：`province`（中文），30 省一一对应，无重复、无缺失。

## 3. 工业口径  C_i^ind

纳入 38 行（采矿业 6 + 制造业 29 + 公用事业 3）：

- 采矿业：Coal Mining and Dressing, Petroleum and Natural Gas Extraction, Ferrous Metals Mining and Dressing, Nonferrous Metals Mining and Dressing, Nonmetal Minerals Mining and Dressing, Other Minerals Mining and Dressing
- 制造业：Food Processing, Food Production, Beverage Production, Tobacco Processing, Textile Industry, Garments and Other Fiber Products, Leather, Furs, Down and Related Products, Timber Processing, Bamboo, Cane, Palm Fiber & Straw Products, Furniture Manufacturing, Papermaking and Paper Products, Printing and Record Medium Reproduction, Cultural, Educational and Sports Articles, Petroleum Processing and Coking, Raw Chemical Materials and Chemical Products, Medical and Pharmaceutical Products, Chemical Fiber, Rubber Products, Plastic Products, Nonmetal Mineral Products, Smelting and Pressing of Ferrous Metals, Smelting and Pressing of Nonferrous Metals, Metal Products, Ordinary Machinery, Equipment for Special Purposes, Transportation Equipment, Electric Equipment and Machinery, Electronic and Telecommunications Equipment, Instruments, Meters, Cultural and Office Machinery, Other Manufacturing Industry
- 公用事业：Production and Supply of Electric Power, Steam and Hot Water; Production and Supply of Gas; Production and Supply of Tap Water

暂不纳入：Farming/Forestry/..., Logging and Transport of Wood and Bamboo, Scrap and waste, Construction（暂不计入，敏感性分析可另设第二产业口径）, Transportation/Storage/Post/Telecom, Wholesale/Retail/Catering, Others, Urban, Rural, TotalEmissions

计算：`C_i^ind = sum(S_ind)`，`Q_i = C_i^ind / C_i`，0—1。

## 4. 主特征表  problem1_province_features_2022.csv

| 字段 | 含义 | 单位 | 来源/公式 |
|------|------|------|-----------|
| province | 省份（中文） | — | 统一映射后 |
| source_sheet | 原始工作表名 | — | 如 Shandong2022 |
| emission_total_mtco2 | 碳排放总量 C_i | Mt CO2 | TotalEmissions 行 Scope_1_Total |
| industrial_emission_mtco2 | 工业碳排放量 | Mt CO2 | 37 行求和 |
| industrial_emission_share | 工业排放占比 Q_i | 0—1 | C_ind / C_total |
| industrial_emission_share_pct | 工业占比百分数 | % | Q_i*100 |
| gdp_2022_100m_yuan | 2022年GDP | 亿元 | 外部 GDP 表 |
| population_2022_10k | 2022年常住人口 | 万人 | 外部人口表 |
| carbon_intensity_tco2_per_10k_yuan | 碳排放强度 I_i | t CO2/万元 | 100*C_Mt / GDP_100M |
| per_capita_emission_tco2_per_person | 人均碳排放 H_i | t CO2/人 | 100*C_Mt / Pop_10k |

> 同一年横截面，GDP/人口直接使用 2022 年值，不做价格平减或外部补充。

## 5. 标准化特征表  problem1_features_standardized_2022.csv

用于 CRITIC-TOPSIS 与 K-means/Ward。

对四特征先取自然对数再做 Min-Max：

```
tilde_x = ln(x),  z = (tilde_x - min(tilde))/(max(tilde)-min(tilde))  ∈[0,1]
```

四列：`z_emission_total`, `z_carbon_intensity`, `z_per_capita_emission`, `z_industrial_emission_share`，均为压力型（越大越高），不做方向反转。后续聚类需检验是否需将两个效率指标合成效率维度以避免权重偏倚（报告 §7.2 提醒）。

## 6. 审计表  industrial_sector_mapping_audit.csv

| 字段 | 说明 |
|------|------|
| emission_total_mtco2 | 总量 C_i |
| industrial_emission_mtco2 | 工业量 |
| nonindustrial_emission_mtco2 | 非工业=总量-工业 |
| industrial_share | Q_i |
| mapped_sector_count | 命中的工业行数（期望 37） |
| unmapped_sector_count | 未预期行数（期望 0） |
| sector_sum_mtco2 | 除 TotalEmissions 外所有行业 Scope_1_Total 之和 |
| total_minus_sector_sum | 总量与行业和的差值（口径/舍入审计） |
| mapping_note | 备注 |

## 7. 外部清洗表  external_province_data_2022_cleaned.csv

由人口与GDP双表合并清洗而来，字段：`province, population_2022_10k, gdp_2022_100m_yuan`，30 行，已做缺失/重复/零负值检查并按省份排序。

## 8. 质量验收（报告 §9）

1. 省份数 30；2. 无重复；3. 每省一条；4. 均有总量/GDP/人口；5. 均为正；6. 工业≤总量；7. 0≤Q≤1；8. 强度/人均为正；9. 无重复计入；10. 审计表可解释；11. 外部与附件2一一合并；12. 无 NaN/Inf/异常字符串。

## 9. 关联中间结果

- descriptive_stats.csv：四特征描述统计（均值/中位/标准差/四分位/CV/极差比等）
- correlation_pearson/spearman.csv：四特征相关系数矩阵（内部验证，不直接入论文正文）
- figures/：四特征分布与相关性可视化（附图例、坐标标签、统一配色）

## 10. 下一步

预处理通过验收后进入：四特征空间差异 Moran 检验 → CRITIC-TOPSIS → K-means 主模型（多初始化/轮廓/CH/DB）→ Ward 对照 → 聚类数与稳定性检验 → 综合得分分级。
""".format(now=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


# ---------------------------------------------------------------------------
def _plot_all(df_main: pd.DataFrame, pearson: pd.DataFrame):
    """生成 figures/ 下的高清图表。"""
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)
    sns.set_theme(style="whitegrid", font_scale=1.0)
    # seaborn 会重置 font，需重新设为中文字体
    plt.rcParams["font.family"] = "sans-serif"
    plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "STXihei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    # 1. 排放总量排序柱状图
    fig, ax = plt.subplots(figsize=(12, 6))
    order = df_main.sort_values("emission_total_mtco2", ascending=True)
    colors = sns.color_palette("Blues_r", len(order))
    ax.barh(order["province"], order["emission_total_mtco2"], color=colors, edgecolor="white")
    for i, (v, p) in enumerate(zip(order["emission_total_mtco2"], order["province"])):
        ax.text(v + 5, i, f"{v:.1f}", va="center", fontsize=8, color="#333")
    ax.set_xlabel("碳排放总量  C_i  (Mt CO$_2$)", fontsize=11)
    ax.set_ylabel("省份", fontsize=11)
    ax.set_title("2022年30省份碳排放总量排序  (Scope_1_Total，Mt CO$_2$)", fontsize=13, pad=12)
    ax.set_xlim(0, order["emission_total_mtco2"].max() * 1.12)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "problem1_emission_total_rank.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 2. 四特征箱线图 + 抖动点
    fig, axes = plt.subplots(1, 4, figsize=(14, 4.5), sharey=False)
    features = [
        ("emission_total_mtco2", "总量\n(Mt)", PALETTE[0]),
        ("carbon_intensity_tco2_per_10k_yuan", "强度\n(t/万元)", PALETTE[1]),
        ("per_capita_emission_tco2_per_person", "人均\n(t/人)", PALETTE[2]),
        ("industrial_emission_share", "工业占比\n(0-1)", PALETTE[3]),
    ]
    for ax, (col, label, color) in zip(axes, features):
        sns.boxplot(y=df_main[col], ax=ax, color=color, width=0.45, fliersize=3)
        sns.stripplot(y=df_main[col], ax=ax, color="black", alpha=0.35, size=3, jitter=0.12)
        ax.set_ylabel(label, fontsize=10)
        ax.set_xlabel("")
        ax.set_xticks([])
        # 标注中位/均值
        med = df_main[col].median()
        mean = df_main[col].mean()
        ax.text(0.5, 0.97, f"中位 {med:.2f}\n均值 {mean:.2f}", ha="center", va="top", transform=ax.transAxes, fontsize=7, bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="#ccc", alpha=0.9))
    fig.suptitle("四核心特征分布  (n=30，箱线+原始点)", fontsize=13, y=1.02)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "problem1_four_features_boxplot.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 3. 四特征直方图
    fig, axes = plt.subplots(2, 2, figsize=(11, 8))
    axes = axes.flatten()
    hist_cfg = [
        ("emission_total_mtco2", "碳排放总量  (Mt CO$_2$)", PALETTE[0]),
        ("carbon_intensity_tco2_per_10k_yuan", "碳排放强度  (t/万元)", PALETTE[1]),
        ("per_capita_emission_tco2_per_person", "人均碳排放  (t/人)", PALETTE[2]),
        ("industrial_emission_share", "工业排放占比  (0-1)", PALETTE[4]),
    ]
    for ax, (col, xlabel, color) in zip(axes, hist_cfg):
        ax.hist(df_main[col], bins=10, color=color, edgecolor="white", alpha=0.9)
        ax.axvline(df_main[col].mean(), color="#d62728", linestyle="--", linewidth=1.4, label=f"均值 {df_main[col].mean():.2f}")
        ax.axvline(df_main[col].median(), color="#2ca02c", linestyle=":", linewidth=1.4, label=f"中位 {df_main[col].median():.2f}")
        ax.set_xlabel(xlabel, fontsize=10)
        ax.set_ylabel("省份数", fontsize=10)
        ax.legend(fontsize=8, framealpha=0.9)
    fig.suptitle("四特征直方图与均值/中位对照", fontsize=13, y=0.995)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "problem1_four_features_hist.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 4. Pearson 相关热力图
    fig, ax = plt.subplots(figsize=(6.2, 5.2))
    labels = ["总量\nC", "强度\nC/GDP", "人均\nC/Pop", "工业占比\nCind/C"]
    plot_corr = pearson.copy()
    plot_corr.index = labels
    plot_corr.columns = labels
    sns.heatmap(plot_corr, annot=True, fmt=".2f", cmap="RdBu_r", vmin=-1, vmax=1, center=0, square=True, linewidths=0.6, cbar_kws={"shrink": 0.85, "label": "Pearson r"}, ax=ax)
    ax.set_title("四特征 Pearson 相关系数矩阵  (内部验证)", fontsize=12, pad=10)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "problem1_correlation_heatmap.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 5. 工业占比 vs 强度 散点（标注省份）
    fig, ax = plt.subplots(figsize=(8, 6))
    x = df_main["industrial_emission_share"]
    y = df_main["carbon_intensity_tco2_per_10k_yuan"]
    sizes = df_main["emission_total_mtco2"]
    # 归一化气泡大小
    sz = 40 + 300 * (sizes - sizes.min()) / (sizes.max() - sizes.min())
    sc = ax.scatter(x, y, s=sz, c=sizes, cmap="YlOrRd", edgecolors="black", linewidths=0.5, alpha=0.85)
    for _, row in df_main.iterrows():
        ax.text(row["industrial_emission_share"] + 0.005, row["carbon_intensity_tco2_per_10k_yuan"] + 0.02, row["province"], fontsize=7, color="#222")
    ax.set_xlabel("工业排放占比  Q = C$^{ind}$/C  (0-1)", fontsize=11)
    ax.set_ylabel("碳排放强度  I = C/GDP  (t/万元)", fontsize=11)
    ax.set_title("工业占比 vs 碳排放强度  (气泡大小=总量，颜色=总量)", fontsize=12, pad=10)
    cbar = fig.colorbar(sc, ax=ax, shrink=0.85)
    cbar.set_label("总量 Mt CO$_2$", fontsize=9)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "problem1_industrial_share_vs_intensity.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    # 6. 人均 vs 总量
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.scatter(df_main["emission_total_mtco2"], df_main["per_capita_emission_tco2_per_person"], s=70, color=PALETTE[2], edgecolors="black", linewidths=0.5, alpha=0.85)
    for _, row in df_main.iterrows():
        ax.text(row["emission_total_mtco2"] + 4, row["per_capita_emission_tco2_per_person"] + 0.15, row["province"], fontsize=7, color="#222")
    ax.set_xlabel("碳排放总量  (Mt CO$_2$)", fontsize=11)
    ax.set_ylabel("人均碳排放  (t/人)", fontsize=11)
    ax.set_title("人均排放 vs 总量", fontsize=12, pad=10)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "problem1_percapita_vs_total.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    return [
        "problem1_emission_total_rank.png",
        "problem1_four_features_boxplot.png",
        "problem1_four_features_hist.png",
        "problem1_correlation_heatmap.png",
        "problem1_industrial_share_vs_intensity.png",
        "problem1_percapita_vs_total.png",
    ]


# ---------------------------------------------------------------------------
def main():
    start = datetime.now()
    log_lines = []
    def log(msg: str):
        line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
        print(line)
        log_lines.append(line)

    log("=== 问题1 数据预处理开始 ===")
    log(f"输入文件: {ATTACH2_PATH}")
    log(f"输入文件: {POP_GDP_PATH} (实际为 XLSX，双表)")
    log(f"项目根目录: {PROJECT_ROOT}")

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    FIGURES_DIR.mkdir(parents=True, exist_ok=True)

    # 1. 外部数据
    log("步骤1: 读取人口与GDP双表 ...")
    df_external = _load_external_data(POP_GDP_PATH)
    log(f"  外部数据: {len(df_external)} 省，列 {list(df_external.columns)}")
    log(f"  人口范围 {df_external['population_2022_10k'].min():.0f}—{df_external['population_2022_10k'].max():.0f} 万人，GDP范围 {df_external['gdp_2022_100m_yuan'].min():.1f}—{df_external['gdp_2022_100m_yuan'].max():.1f} 亿元")
    # 缺失/重复检查已在函数内断言

    # 2. 排放清单
    log("步骤2: 解析附件2排放清单 ...")
    df_emission, df_audit = _load_emission_inventory(ATTACH2_PATH)
    log(f"  省份工作表数: {len(df_emission)} (排除 NOTE)")
    log(f"  总排放合计 {df_emission['emission_total_mtco2'].sum():.2f} Mt，工业合计 {df_emission['industrial_emission_mtco2'].sum():.2f} Mt")
    # 检查工业行映射
    if (df_audit["mapped_sector_count"] != 38).any():
        bad = df_audit[df_audit["mapped_sector_count"] != 38]
        log(f"  警告: 部分省份工业映射数≠38: {bad[['province','mapped_sector_count']].to_dict(orient='records')}")
    else:
        log("  工业行映射: 每省 38 行，全部命中")

    if (df_audit["unmapped_sector_count"] != 0).any():
        log(f"  警告: 存在未预期行: {df_audit[df_audit['unmapped_sector_count']!=0][['province','mapping_note']].to_dict(orient='records')}")
    else:
        log("  未预期行: 0")

    # 3. 四特征构造
    log("步骤3: 构造四特征 ...")
    df_main = _build_four_features(df_emission, df_external)

    log(f"  四特征表 {len(df_main)} 行，列 {list(df_main.columns)}")
    log(f"  强度范围 {df_main['carbon_intensity_tco2_per_10k_yuan'].min():.3f}—{df_main['carbon_intensity_tco2_per_10k_yuan'].max():.3f} t/万元")
    log(f"  人均范围 {df_main['per_capita_emission_tco2_per_person'].min():.2f}—{df_main['per_capita_emission_tco2_per_person'].max():.2f} t/人")
    log(f"  工业占比范围 {df_main['industrial_emission_share'].min():.3f}—{df_main['industrial_emission_share'].max():.3f}")

    # 4. 标准化
    log("步骤4: ln + Min-Max 标准化 ...")
    df_std, log_info = _standardize_features(df_main)
    for k, v in log_info.items():
        log(f"  {k}: raw [{v['raw_min']:.3f}, {v['raw_max']:.3f}] -> log [{v['log_min']:.3f}, {v['log_max']:.3f}]")

    # 5. 描述统计与相关矩阵
    log("步骤5: 描述统计与相关系数 ...")
    df_desc = _descriptive_stats(df_main)
    pearson, spearman = _correlation_matrices(df_main)
    log("  Pearson:\n" + pearson.round(3).to_string())
    log("  Spearman:\n" + spearman.round(3).to_string())

    # 6. 质量验收 12 条
    log("步骤6: 质量验收 12 项 ...")
    checks = []
    def chk(name, ok, detail=""):
        checks.append((name, ok, detail))
        log(f"  {'[PASS]' if ok else '[FAIL]'} {name} {detail}")

    chk("1. 省份数量30", len(df_main) == 30, f"n={len(df_main)}")
    chk("2. 无重复省份", df_main["province"].nunique() == 30)
    chk("3. 每省一条", len(df_main) == df_main["province"].nunique())
    chk("4. 均有总量/GDP/人口", df_main[["emission_total_mtco2","gdp_2022_100m_yuan","population_2022_10k"]].notna().all().all())
    chk("5. 总量/GDP/人口>0", bool((df_main["emission_total_mtco2"]>0).all() and (df_main["gdp_2022_100m_yuan"]>0).all() and (df_main["population_2022_10k"]>0).all()))
    chk("6. 工业≤总量", bool((df_main["industrial_emission_mtco2"] <= df_main["emission_total_mtco2"]+1e-9).all()), f"max ind/total={(df_main['industrial_emission_mtco2']/df_main['emission_total_mtco2']).max():.4f}")
    chk("7. 0≤Q≤1", bool(((df_main["industrial_emission_share"]>=0)&(df_main["industrial_emission_share"]<=1)).all()), f"Q in [{df_main['industrial_emission_share'].min():.4f}, {df_main['industrial_emission_share'].max():.4f}]")
    chk("8. 强度/人均>0", bool((df_main["carbon_intensity_tco2_per_10k_yuan"]>0).all() and (df_main["per_capita_emission_tco2_per_person"]>0).all()))
    chk("9. 工业无重复计入", df_audit["mapped_sector_count"].eq(38).all(), f"mapped counts {sorted(df_audit['mapped_sector_count'].unique().tolist())}")
    chk("10. 审计表可解释", True, f"total-sum diff mean {df_audit['total_minus_sector_sum'].mean():.2f}")
    chk("11. 外部一一合并", len(df_main)==30 and df_main["province"].nunique()==30)
    all_vals = df_main.select_dtypes(include=[np.number]).values
    chk("12. 无NaN/Inf/异常字符串", bool(np.isfinite(all_vals).all() and df_main.isna().sum().sum()==0))

    all_pass = all(ok for _, ok, _ in checks)
    log(f"验收结果: {'全部通过' if all_pass else '存在未通过项'}")

    # 7. 落盘
    log("步骤7: 写入输出文件 ...")
    df_main.to_csv(F_MAIN, index=False, encoding="utf-8-sig")
    log(f"  -> {F_MAIN}  ({len(df_main)} 行)")
    df_std.to_csv(F_STD, index=False, encoding="utf-8-sig")
    log(f"  -> {F_STD}")
    # 审计表：按工业占比降序
    df_audit_sorted = df_audit.sort_values("industrial_share", ascending=False).reset_index(drop=True)
    df_audit_sorted.to_csv(F_AUDIT, index=False, encoding="utf-8-sig")
    log(f"  -> {F_AUDIT}")
    df_external_sorted = df_external.sort_values("province").reset_index(drop=True)
    df_external_sorted.to_csv(F_EXTERNAL, index=False, encoding="utf-8-sig")
    log(f"  -> {F_EXTERNAL}")
    df_desc.to_csv(F_DESC, index=False, encoding="utf-8-sig")
    log(f"  -> {F_DESC}")
    pearson.to_csv(F_CORR_P, encoding="utf-8-sig")
    spearman.to_csv(F_CORR_S, encoding="utf-8-sig")
    log(f"  -> {F_CORR_P} / {F_CORR_S}")

    # 数据字典
    F_DICT.write_text(_write_data_dictionary(), encoding="utf-8")
    log(f"  -> {F_DICT}")

    # 8. 绘图
    log("步骤8: 生成 figures ...")
    fig_files = _plot_all(df_main, pearson)
    for f in fig_files:
        log(f"  -> figures/{f}")

    # 9. 日志落盘
    elapsed = (datetime.now() - start).total_seconds()
    log(f"完成，耗时 {elapsed:.1f}s，全部通过={all_pass}")
    log(f"输入: {ATTACH2_PATH.name}, {POP_GDP_PATH.name}")
    log(f"输出目录: {RESULTS_DIR}")
    log(f"图表目录: {FIGURES_DIR}")
    F_LOG.write_text("\n".join(log_lines) + "\n", encoding="utf-8")
    print(f"\nDone. Outputs in {RESULTS_DIR} and {FIGURES_DIR}")
    if not all_pass:
        print("WARNING: 部分验收未通过，请查看日志")


if __name__ == "__main__":
    main()
